"""Resource-bounded, no-database XLSX parser for branch imports."""
from __future__ import annotations

import datetime as dt
import json
from pathlib import Path
import platform
import sys


BRANCH_HEADERS = [
    "门店编号*", "店名*", "区域*", "省", "市", "区/县", "详细地址*", "店长姓名", "店长工号",
    "店长手机号", "门店类型", "开业日期", "营业面积㎡", "座位数/房间数/工位数",
    "经度", "纬度", "启用状态*", "备注",
]
VALUE_HEADERS = [
    "门店编号*", "指标编码*", "期间开始*", "期间结束*", "数值*",
    "单位*", "数据来源*", "备注",
]
SHEET_NAMES = ["门店主表", "经营数据", "填写说明", "示例（不要导入）"]
MAX_SHEETS = 4
# One workbook may carry the whole company master plus a separate operating
# fact table. Keep independent sheet limits so optional business data never
# steals capacity from the promised 20,000-store master-data contract.
MAX_BRANCH_DATA_ROWS = 20_000
MAX_BUSINESS_DATA_ROWS = 40_000
MAX_CELLS = 700_000
# Excel itself caps ordinary cell text at 32,767 characters, but every import
# field is contractually at most 500 characters.  A modest parser-side ceiling
# stops oversized inline/shared strings before JSON projection, while retaining
# ample headroom for headings and forward-compatible template notes.
MAX_CELL_TEXT_CHARS = 4_096
# The standard 20k-store + 40k-fact workbook is about 3.9M text characters.
# Eight Mi characters preserves that full row capacity while bounding the sum
# of otherwise individually-small strings before the worker emits JSON.
MAX_TOTAL_TEXT_CHARS = 8 * 1024 * 1024


class ParseError(ValueError):
    def __init__(self, code: str):
        super().__init__(code)
        self.code = code


def _apply_resource_limits() -> bool:
    try:
        import resource
    except ImportError:
        return False
    # A worst-case accepted workbook is 60k data rows / 680k cells. openpyxl's
    # read-only model plus the JSON projection stays comfortably below this
    # ceiling while malformed archives remain bounded by the parent preflight.
    memory = 768 * 1024 * 1024
    applied = 0
    for name, limits in (
            ("RLIMIT_CORE", (0, 0)),
            ("RLIMIT_AS", (memory, memory)),
            ("RLIMIT_CPU", (55, 60)),
            ("RLIMIT_FSIZE", (96 * 1024 * 1024, 96 * 1024 * 1024)),
            ("RLIMIT_NOFILE", (48, 48)),
            ("RLIMIT_NPROC", (16, 16)),
    ):
        try:
            key = getattr(resource, name, None)
            if key is not None:
                resource.setrlimit(key, limits)
                applied += 1
        except (OSError, ValueError):
            # macOS exposes RLIMIT_AS but its kernel rejects finite values for
            # framework Python ("current limit exceeds maximum"). The worker
            # still has CPU/file/fd/process limits plus a parent timeout and
            # preflight-expanded-size cap. Linux must accept the address cap.
            if name != "RLIMIT_AS" or platform.system() != "Darwin":
                return False
    return applied >= 5


def _json_value(value):
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _read_sheet(
    sheet,
    headers: list[str],
    counters: dict[str, int],
    *,
    max_data_rows: int,
) -> list[dict]:
    declared_rows = int(sheet.max_row or 0)
    declared_columns = int(sheet.max_column or 0)
    if (
        declared_rows > max_data_rows + 3
        or declared_columns > len(headers)
    ):
        raise ParseError("ROW_LIMIT_EXCEEDED")
    rows = sheet.iter_rows(values_only=False)
    output: list[dict] = []
    saw_header = False
    for row_number, cells_in_row in enumerate(rows, 1):
        if row_number > max_data_rows + 3:
            raise ParseError("ROW_LIMIT_EXCEEDED")
        counters["rows"] += 1
        counters["cells"] += len(cells_in_row)
        if counters["cells"] > MAX_CELLS:
            raise ParseError("CELL_LIMIT_EXCEEDED")
        for cell in cells_in_row:
            if cell.data_type == "f" or (
                isinstance(cell.value, str) and cell.value.startswith("=")
            ):
                raise ParseError("XLSX_FORMULA_FORBIDDEN")
            if isinstance(cell.value, str):
                text_chars = len(cell.value)
                if text_chars > MAX_CELL_TEXT_CHARS:
                    raise ParseError("CELL_TEXT_LIMIT_EXCEEDED")
                counters["text_chars"] += text_chars
                if counters["text_chars"] > MAX_TOTAL_TEXT_CHARS:
                    raise ParseError("TEXT_BUDGET_EXCEEDED")
        if row_number == 3:
            actual = [str(cell.value or "").strip() for cell in cells_in_row]
            if actual != headers:
                raise ParseError("HEADER_INVALID")
            saw_header = True
            continue
        if row_number < 4:
            continue
        values = [_json_value(cell.value) for cell in cells_in_row]
        values.extend([None] * (len(headers) - len(values)))
        if all(value is None or str(value).strip() == "" for value in values):
            continue
        output.append({
            "row_number": row_number,
            "data": dict(zip(headers, values[:len(headers)])),
        })
    if not saw_header:
        raise ParseError("HEADER_INVALID")
    return output


def parse(path: Path) -> dict:
    from openpyxl import load_workbook

    workbook = load_workbook(
        str(path), read_only=True, data_only=False, keep_links=False,
    )
    try:
        if len(workbook.worksheets) != MAX_SHEETS:
            raise ParseError("SHEET_LIMIT_EXCEEDED")
        names = workbook.sheetnames
        if names != SHEET_NAMES:
            raise ParseError("HEADER_INVALID")
        counters = {"rows": 0, "cells": 0, "text_chars": 0}
        branches = _read_sheet(
            workbook.worksheets[0], BRANCH_HEADERS, counters,
            max_data_rows=MAX_BRANCH_DATA_ROWS,
        )
        values = _read_sheet(
            workbook.worksheets[1], VALUE_HEADERS, counters,
            max_data_rows=MAX_BUSINESS_DATA_ROWS,
        )
        return {"branches": branches, "business_values": values}
    finally:
        workbook.close()


def main(argv: list[str]) -> int:
    if len(argv) != 2 or not _apply_resource_limits():
        return 2
    try:
        result = parse(Path(argv[1]))
        sys.stdout.write(json.dumps({"ok": True, **result}, ensure_ascii=False))
        return 0
    except ParseError as exc:
        sys.stdout.write(json.dumps({"ok": False, "error_code": exc.code}))
        return 0
    except Exception:
        sys.stdout.write(json.dumps({"ok": False, "error_code": "XLSX_INVALID"}))
        return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
