"""Schema 52 branch-master XLSX import contracts."""
from __future__ import annotations

import io
import json
import os
import re
import sys
import tempfile
import threading
import time
import unittest
import zipfile
from pathlib import Path
from unittest import mock

from openpyxl import Workbook

from app import auth, db, inspectionimport, inspectionstandards


BRANCH_HEADERS = [
    "门店编号*", "店名*", "区域*", "省", "市", "区/县", "详细地址*", "店长姓名", "店长工号",
    "店长手机号", "门店类型", "开业日期", "营业面积㎡", "座位数/房间数/工位数",
    "经度", "纬度", "启用状态*", "备注",
]
VALUE_HEADERS = ["门店编号*", "指标编码*", "期间开始*", "期间结束*", "数值*", "单位*", "数据来源*", "备注"]


def branch_row(code="S001", name="一号店", phone="13800138000"):
    return [code, name, "华东", "浙江省", "杭州市", "西湖区", "文三路1号", "张三", "E1001",
            phone, "旗舰店", "2026-01-02", 120.5, 60, 120.123, 30.234, "启用", "测试"]


def xlsx(rows, values=None, formula=False):
    book = Workbook()
    sheet = book.active
    sheet.title = "门店主表"
    sheet.append(["派活 · 门店批量导入模板"])
    sheet.append(["填写说明"])
    sheet.append(BRANCH_HEADERS)
    for row in rows:
        sheet.append(row)
    if formula:
        sheet["B4"] = '="恶意门店"'
    value_sheet = book.create_sheet("经营数据")
    value_sheet.append(["派活 · 门店经营数据（选填）"])
    value_sheet.append(["填写说明"])
    value_sheet.append(VALUE_HEADERS)
    if values is not None:
        for row in values:
            value_sheet.append(row)
    book.create_sheet("填写说明")
    book.create_sheet("示例（不要导入）")
    out = io.BytesIO()
    book.save(out)
    return out.getvalue()


def rewritten_xlsx(data: bytes, replacements: dict[str, bytes]) -> bytes:
    """Replace OOXML members without creating ambiguous duplicate ZIP names."""
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(data)) as source, zipfile.ZipFile(
        output, "w", zipfile.ZIP_DEFLATED
    ) as target:
        for item in source.infolist():
            if item.filename not in replacements:
                target.writestr(item, source.read(item.filename))
        for name, body in replacements.items():
            target.writestr(name, body)
    return output.getvalue()


class InspectionBranchImportCase(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.old_path = db.DB_PATH
        db._close_all_connections()
        db._conn = db._conn_path = None
        db.DB_PATH = os.path.join(self.tmp.name, "import.db")
        db.conn()
        db.insert("tenants", {"id": 2, "name": "甲企业"})
        db.execute("INSERT INTO tenant_industry(tenant_id,industry_key,is_primary,created_at) VALUES(2,'restaurant',1,0)")
        db.insert("users", {
            "id": 20, "tenant_id": 2, "username": "owner-import",
            "password_hash": "test", "role": "owner", "modules_json": "[]",
            "enabled": 1,
        })

    def tearDown(self):
        db._close_all_connections()
        db._conn = db._conn_path = None
        db.DB_PATH = self.old_path
        self.tmp.cleanup()

    def test_schema52_contract_and_legacy_nullable_code(self):
        columns = {row["name"] for row in db.q("PRAGMA table_info(store_branch)")}
        self.assertTrue({"store_code", "province", "city", "district", "manager_name",
                         "manager_employee_no", "manager_phone", "store_type", "opened_on",
                         "area_sqm", "seat_count", "longitude", "latitude", "remark",
                         "row_version"} <= columns)
        tables = {row["name"] for row in db.q("SELECT name FROM sqlite_master WHERE type='table'")}
        self.assertTrue({"inspection_branch_import", "inspection_branch_import_row",
                         "inspection_business_value"} <= tables)
        import_columns = {
            row["name"]
            for row in db.q("PRAGMA table_info(inspection_branch_import)")
        }
        self.assertTrue(
            {
                "catalog_version", "catalog_sha256",
                "business_create_count", "business_update_count",
                "business_skip_count", "business_error_count",
                "staging_purged_at",
            } <= import_columns
        )
        import_row_columns = {
            row["name"]
            for row in db.q("PRAGMA table_info(inspection_branch_import_row)")
        }
        self.assertTrue({
            "existing_branch_id", "existing_row_version",
            "existing_business_value_id", "existing_business_row_version",
        } <= import_row_columns)
        business_columns = {
            row["name"]
            for row in db.q("PRAGMA table_info(inspection_business_value)")
        }
        self.assertIn("row_version", business_columns)
        indexes = {row["name"] for row in db.q("SELECT name FROM sqlite_master WHERE type='index'")}
        self.assertTrue({"idx_store_branch_code", "idx_inspection_branch_import_request",
                         "idx_inspection_branch_import_source", "idx_inspection_import_row",
                         "idx_inspection_business_value_natural",
                         "idx_inspection_business_value_period"} <= indexes)
        name_index = next(
            row for row in db.q("PRAGMA index_list(store_branch)")
            if row["name"] == "idx_store_branch_name"
        )
        self.assertEqual(0, name_index["unique"])
        legacy = db.insert("store_branch", {"tenant_id": 2, "industry_key": "restaurant", "name": "旧店"})
        self.assertIsNone(db.one("SELECT store_code FROM store_branch WHERE id=?", (legacy,))["store_code"])
        self.assertEqual(
            db.LATEST_SCHEMA_VERSION,
            db.one("PRAGMA user_version")["user_version"],
        )
        self.assertEqual("inspection-branch-master-import", db.one("SELECT name FROM schema_version WHERE version=52")["name"])

    def test_preview_has_no_business_side_effect_masks_pii_then_commit_replays(self):
        data = xlsx([branch_row()], [["S001", "common.net_revenue", "2026-01-01", "2026-01-31", 1234.5, "CNY", "erp", ""]])
        preview = inspectionimport.preview_import(2, 20, "restaurant", "branch-preview-0001", "branches.xlsx", data)
        self.assertEqual(
            inspectionstandards.CATALOG_VERSION,
            preview["catalog_version"],
        )
        self.assertEqual(
            inspectionstandards.version_summary("restaurant")["sha256"],
            preview["catalog_sha256"],
        )
        self.assertEqual({"create": 1, "update": 0, "skip": 0, "error": 0}, preview["counts"])
        rendered = json.dumps(preview, ensure_ascii=False)
        self.assertNotIn("13800138000", rendered)
        self.assertNotIn("E1001", rendered)
        self.assertIn("138****8000", rendered)
        # A clean preview is still committable, but its private staging copy is
        # authenticated ciphertext from the first SQLite write onward.
        clean_stage = db.one(
            "SELECT payload_json,masked_payload_json "
            "FROM inspection_branch_import_row WHERE import_id=?",
            (preview["import_id"],),
        )
        self.assertTrue(clean_stage["payload_json"].startswith(
            inspectionimport._STAGING_ENCRYPTED_PREFIX
        ))
        self.assertNotIn("13800138000", clean_stage["payload_json"])
        self.assertNotIn("13800138000", clean_stage["masked_payload_json"])
        encrypted_business = db.one(
            "SELECT business_values_json FROM inspection_branch_import "
            "WHERE id=?", (preview["import_id"],),
        )["business_values_json"]
        self.assertTrue(encrypted_business.startswith(
            inspectionimport._STAGING_ENCRYPTED_PREFIX
        ))
        self.assertNotIn(
            "1234.5", encrypted_business,
        )
        self.assertEqual(0, db.one("SELECT COUNT(*) n FROM store_branch")["n"])
        self.assertEqual(0, db.one("SELECT COUNT(*) n FROM inspection_business_value")["n"])
        committed = inspectionimport.commit_import(2, 20, preview["import_id"], "restaurant")
        self.assertEqual("committed", committed["status"])
        self.assertEqual(1, db.one("SELECT COUNT(*) n FROM store_branch")["n"])
        self.assertEqual(1, db.one("SELECT COUNT(*) n FROM inspection_business_value")["n"])
        ledger = db.one(
            "SELECT business_values_json FROM inspection_branch_import "
            "WHERE id=?", (preview["import_id"],),
        )
        self.assertEqual("[]", ledger["business_values_json"])
        staged_payloads = " ".join(
            row["payload_json"] for row in db.q(
                "SELECT payload_json FROM inspection_branch_import_row "
                "WHERE import_id=?", (preview["import_id"],),
            )
        )
        self.assertNotIn("13800138000", staged_payloads)
        self.assertNotIn("E1001", staged_payloads)
        self.assertEqual(committed["import_id"], inspectionimport.commit_import(2, 20, preview["import_id"], "restaurant")["import_id"])
        self.assertEqual(committed["import_id"], inspectionimport.preview_import(
            2, 20, "restaurant", "branch-preview-0001", "branches.xlsx", data)["import_id"])
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "REQUEST_KEY_CONFLICT"):
            inspectionimport.preview_import(2, 20, "restaurant", "branch-preview-0001", "changed.xlsx", xlsx([branch_row(name="篡改")]))

        revised = xlsx([branch_row()], [[
            "S001", "common.net_revenue", "2026-01-01", "2026-01-31",
            1400, "CNY", "audited-erp", "补录",
        ]])
        revised_preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-preview-0002", "revised.xlsx", revised
        )
        inspectionimport.commit_import(
            2, 20, revised_preview["import_id"], "restaurant"
        )
        self.assertEqual(
            1, db.one("SELECT COUNT(*) n FROM inspection_business_value")["n"]
        )
        revised_value = db.one("SELECT * FROM inspection_business_value")
        self.assertEqual((1400, "audited-erp", "补录"), (
            revised_value["value"], revised_value["source_ref"],
            revised_value["remark"],
        ))

    def test_store_code_drives_update_skip_and_business_values(self):
        existing = db.insert("store_branch", {"tenant_id": 2, "industry_key": "restaurant",
                                               "store_code": "S001", "name": "旧名", "manager_phone": "13900000000"})
        data = xlsx([branch_row(name="新名"), branch_row("S002", "二号店")],
                    [["S001", "common.transactions", "2026-02-01", "2026-02-28", 88, "count", "pos", ""]])
        preview = inspectionimport.preview_import(2, 20, "restaurant", "branch-update-0001", "master.xlsx", data)
        self.assertEqual((1, 1), (preview["counts"]["update"], preview["counts"]["create"]))
        inspectionimport.commit_import(2, 20, preview["import_id"], "restaurant")
        updated = db.one("SELECT * FROM store_branch WHERE id=?", (existing,))
        self.assertEqual(("新名", "13800138000"), (updated["name"], updated["manager_phone"]))
        self.assertEqual(existing, db.one("SELECT branch_id FROM inspection_business_value")["branch_id"])
        same = inspectionimport.preview_import(2, 20, "restaurant", "branch-skip-0001", "same.xlsx", data)
        self.assertEqual(2, same["counts"]["skip"])

    def test_successful_business_rows_are_visible_counted_and_committed(self):
        data = xlsx([branch_row()], [
            [
                "S001", "common.net_revenue", "2026-04-01", "2026-04-30",
                4321.5, "CNY", "erp-revenue", "月结",
            ],
            [
                "S001", "common.transactions", "2026-04-01", "2026-04-30",
                321, "count", "pos-orders", "月结",
            ],
        ])
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-business-visible-0001",
            "business-visible.xlsx", data,
        )
        self.assertEqual(
            {"create": 1, "update": 0, "skip": 0, "error": 0},
            preview["counts"],
        )
        self.assertEqual(
            {"create": 2, "update": 0, "skip": 0, "error": 0},
            preview["business_counts"],
        )
        self.assertEqual(2, preview["business_total_rows"])
        business_rows = [
            row for row in preview["rows"] if row["row_kind"] == "business"
        ]
        self.assertEqual(2, len(business_rows))
        self.assertEqual(
            {"common.net_revenue", "common.transactions"},
            {row["data"]["metric_key"] for row in business_rows},
        )
        revenue = next(
            row for row in business_rows
            if row["data"]["metric_key"] == "common.net_revenue"
        )
        self.assertEqual((4321.5, "CNY", "erp-revenue"), (
            revenue["data"]["value"], revenue["data"]["unit"],
            revenue["data"]["source_ref"],
        ))
        inspectionimport.commit_import(
            2, 20, preview["import_id"], "restaurant"
        )
        self.assertEqual(
            2, db.one("SELECT COUNT(*) n FROM inspection_business_value")["n"]
        )

    def test_reverse_business_value_commit_uses_natural_key_cas(self):
        base = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-business-base-0001", "base.xlsx",
            xlsx([branch_row("BVCAS001", "经营CAS店")]),
        )
        inspectionimport.commit_import(
            2, 20, base["import_id"], "restaurant"
        )
        branch = db.one(
            "SELECT id,row_version FROM store_branch WHERE store_code='BVCAS001'"
        )
        value_id = db.insert("inspection_business_value", {
            "tenant_id": 2, "industry_key": "restaurant",
            "branch_id": branch["id"], "import_id": base["import_id"],
            "metric_key": "common.net_revenue",
            "period_start": "2026-05-01", "period_end": "2026-05-31",
            "value": 50, "unit": "CNY", "source_ref": "opening-ledger",
        })
        def business_preview(request_key, value, data=None):
            return inspectionimport.preview_import(
                2, 20, "restaurant", request_key, f"{value}.xlsx",
                data or xlsx([branch_row("BVCAS001", "经营CAS店")], [[
                    "BVCAS001", "common.net_revenue", "2026-05-01",
                    "2026-05-31", value, "CNY", f"ledger-{value}", "",
                ]]),
            )

        first_data = xlsx([branch_row("BVCAS001", "经营CAS店")], [[
            "BVCAS001", "common.net_revenue", "2026-05-01",
            "2026-05-31", 100, "CNY", "ledger-100", "",
        ]])
        first = business_preview(
            "branch-business-cas-first-0001", 100, first_data,
        )
        second = business_preview("branch-business-cas-second-0001", 200)
        self.assertEqual(1, first["counts"]["skip"])
        self.assertEqual(1, second["counts"]["skip"])
        self.assertEqual(1, first["business_counts"]["update"])
        frozen = db.q(
            "SELECT import_id,existing_business_value_id,"
            "existing_business_row_version FROM inspection_branch_import_row "
            "WHERE row_number>=100000 ORDER BY import_id"
        )
        self.assertEqual(
            [(first["import_id"], value_id, 1),
             (second["import_id"], value_id, 1)],
            [(row["import_id"], row["existing_business_value_id"],
              row["existing_business_row_version"]) for row in frozen],
        )

        inspectionimport.commit_import(
            2, 20, second["import_id"], "restaurant"
        )
        with self.assertRaisesRegex(
            inspectionimport.ImportContractError, "IMPORT_STATE_CONFLICT"
        ):
            inspectionimport.commit_import(
                2, 20, first["import_id"], "restaurant"
            )
        stored = db.one(
            "SELECT value,source_ref,row_version FROM inspection_business_value "
            "WHERE id=?", (value_id,),
        )
        self.assertEqual((200, "ledger-200", 2), (
            stored["value"], stored["source_ref"], stored["row_version"],
        ))
        self.assertEqual(
            1,
            db.one(
                "SELECT row_version FROM store_branch WHERE id=?",
                (branch["id"],),
            )["row_version"],
        )
        retired = db.one(
            "SELECT status,staging_purged_at FROM inspection_branch_import "
            "WHERE id=?", (first["import_id"],),
        )
        self.assertEqual("expired", retired["status"])
        self.assertIsNone(retired["staging_purged_at"])
        retained = "\n".join(
            row["payload_json"] for row in db.q(
                "SELECT payload_json FROM inspection_branch_import_row "
                "WHERE import_id=?", (first["import_id"],),
            )
        )
        self.assertIn(inspectionimport._STAGING_ENCRYPTED_PREFIX, retained)
        self.assertNotIn("ledger-100", retained)
        with mock.patch.object(
            inspectionimport, "MAX_ACTIVE_PREVIEWS_PER_TENANT", 1,
        ):
            replacement = business_preview(
                "branch-business-cas-retry-0001", 100, first_data,
            )
        self.assertNotEqual(first["import_id"], replacement["import_id"])

    def test_business_rows_keep_each_branch_frozen_version_when_batched(self):
        rows = [
            branch_row("BV001", "经营版本店甲"),
            branch_row("BV002", "经营版本店乙", "13900139000"),
        ]
        base = inspectionimport.preview_import(
            2, 20, "restaurant", "business-branch-versions-base",
            "base.xlsx", xlsx(rows),
        )
        inspectionimport.commit_import(
            2, 20, base["import_id"], "restaurant"
        )
        db.execute(
            "UPDATE store_branch SET row_version=7 WHERE store_code='BV002'"
        )
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "business-branch-versions-page",
            "values.xlsx", xlsx(rows, [
                ["BV001", "common.net_revenue", "2026-06-01", "2026-06-30",
                 101, "CNY", "audited-pos", ""],
                ["BV002", "common.net_revenue", "2026-06-01", "2026-06-30",
                 202, "CNY", "audited-pos", ""],
            ]),
        )
        frozen = db.q(
            "SELECT store_code,existing_row_version "
            "FROM inspection_branch_import_row WHERE import_id=? "
            "AND row_number>=100000 ORDER BY row_number",
            (preview["import_id"],),
        )
        self.assertEqual(
            [("BV001", 1), ("BV002", 7)],
            [(row["store_code"], row["existing_row_version"]) for row in frozen],
        )

    def test_store_code_allows_same_name_for_distinct_branches(self):
        second = branch_row("SAME002", "中心店", "13900139000")
        second[2] = "华南"
        data = xlsx([
            branch_row("SAME001", "中心店"),
            second,
        ])
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-same-name-0001", "same-name.xlsx", data
        )
        self.assertEqual(
            {"create": 2, "update": 0, "skip": 0, "error": 0},
            preview["counts"],
        )
        inspectionimport.commit_import(
            2, 20, preview["import_id"], "restaurant"
        )
        rows = db.q(
            "SELECT store_code,name,row_version FROM store_branch "
            "ORDER BY store_code"
        )
        self.assertEqual(
            [("SAME001", "中心店", 1), ("SAME002", "中心店", 1)],
            [(row["store_code"], row["name"], row["row_version"]) for row in rows],
        )

    def test_reverse_preview_commit_uses_frozen_row_version_cas(self):
        branch_id = db.insert("store_branch", {
            "tenant_id": 2, "industry_key": "restaurant",
            "store_code": "CAS001", "name": "初始店",
        })
        first_data = xlsx([branch_row("CAS001", "第一次修改")])
        first = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-cas-first-0001", "first.xlsx",
            first_data,
        )
        second = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-cas-second-0001", "second.xlsx",
            xlsx([branch_row("CAS001", "第二次修改")]),
        )
        frozen = db.q(
            "SELECT import_id,existing_branch_id,existing_row_version "
            "FROM inspection_branch_import_row ORDER BY import_id"
        )
        self.assertEqual(
            [(first["import_id"], branch_id, 1),
             (second["import_id"], branch_id, 1)],
            [(row["import_id"], row["existing_branch_id"],
              row["existing_row_version"]) for row in frozen],
        )

        inspectionimport.commit_import(
            2, 20, second["import_id"], "restaurant"
        )
        with self.assertRaisesRegex(
            inspectionimport.ImportContractError, "IMPORT_STATE_CONFLICT"
        ):
            inspectionimport.commit_import(
                2, 20, first["import_id"], "restaurant"
            )
        branch = db.one("SELECT * FROM store_branch WHERE id=?", (branch_id,))
        self.assertEqual(("第二次修改", 2), (
            branch["name"], branch["row_version"],
        ))
        self.assertEqual(
            "expired",
            db.one(
                "SELECT status FROM inspection_branch_import WHERE id=?",
                (first["import_id"],),
            )["status"],
        )
        with self.assertRaisesRegex(
            inspectionimport.ImportContractError, "IMPORT_PREVIEW_EXPIRED"
        ):
            inspectionimport.preview_import(
                2, 20, "restaurant", "branch-cas-first-0001", "first.xlsx",
                first_data,
            )
        replacement = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-cas-retry-0001", "retry.xlsx",
            first_data,
        )
        self.assertNotEqual(first["import_id"], replacement["import_id"])

    def test_one_drifted_row_rolls_back_the_entire_batch(self):
        first_id = db.insert("store_branch", {
            "tenant_id": 2, "industry_key": "restaurant",
            "store_code": "BATCH001", "name": "批量原店甲",
        })
        second_id = db.insert("store_branch", {
            "tenant_id": 2, "industry_key": "restaurant",
            "store_code": "BATCH002", "name": "批量原店乙",
        })
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-batch-cas-0001", "batch.xlsx",
            xlsx([
                branch_row("BATCH001", "批量新店甲"),
                branch_row("BATCH002", "批量新店乙", "13900139000"),
            ]),
        )
        db.execute(
            "UPDATE store_branch SET name='外部先改',row_version=row_version+1 "
            "WHERE id=?", (second_id,),
        )
        with self.assertRaisesRegex(
            inspectionimport.ImportContractError, "IMPORT_STATE_CONFLICT"
        ):
            inspectionimport.commit_import(
                2, 20, preview["import_id"], "restaurant"
            )
        first = db.one(
            "SELECT name,row_version FROM store_branch WHERE id=?", (first_id,)
        )
        second_row = db.one(
            "SELECT name,row_version FROM store_branch WHERE id=?", (second_id,)
        )
        self.assertEqual(("批量原店甲", 1), (
            first["name"], first["row_version"],
        ))
        self.assertEqual(("外部先改", 2), (
            second_row["name"], second_row["row_version"],
        ))

    def test_concurrent_create_previews_have_one_atomic_winner(self):
        first = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-race-first-0001", "race-a.xlsx",
            xlsx([branch_row("RACE001", "并发甲")]),
        )
        second = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-race-second-0001", "race-b.xlsx",
            xlsx([branch_row("RACE001", "并发乙")]),
        )
        baseline = db.q(
            "SELECT existing_branch_id,existing_row_version "
            "FROM inspection_branch_import_row ORDER BY import_id"
        )
        self.assertEqual([(0, 0), (0, 0)], [
            (row["existing_branch_id"], row["existing_row_version"])
            for row in baseline
        ])
        gate = threading.Barrier(3)
        outcomes = []
        guard = threading.Lock()

        def submit(import_id):
            gate.wait()
            try:
                inspectionimport.commit_import(
                    2, 20, import_id, "restaurant"
                )
                outcome = "committed"
            except inspectionimport.ImportContractError as exc:
                outcome = exc.code
            finally:
                db._close_thread_connection()
            with guard:
                outcomes.append(outcome)

        threads = [
            threading.Thread(target=submit, args=(item["import_id"],))
            for item in (first, second)
        ]
        for thread in threads:
            thread.start()
        gate.wait()
        for thread in threads:
            thread.join(10)
            self.assertFalse(thread.is_alive())
        self.assertCountEqual(
            ["committed", "IMPORT_STATE_CONFLICT"], outcomes
        )
        self.assertEqual(
            1,
            db.one(
                "SELECT COUNT(*) n FROM store_branch WHERE store_code='RACE001'"
            )["n"],
        )

    def test_frozen_baseline_survives_database_restart(self):
        branch_id = db.insert("store_branch", {
            "tenant_id": 2, "industry_key": "restaurant",
            "store_code": "RESTART001", "name": "重启前",
        })
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-restart-0001", "restart.xlsx",
            xlsx([branch_row("RESTART001", "重启后")]),
        )
        db._close_all_connections()
        db._conn = None
        db._conn_path = None
        db.conn()
        inspectionimport.commit_import(
            2, 20, preview["import_id"], "restaurant"
        )
        branch = db.one("SELECT name,row_version FROM store_branch WHERE id=?", (branch_id,))
        self.assertEqual(("重启后", 2), (
            branch["name"], branch["row_version"],
        ))

    def test_seven_digit_phone_is_irreversibly_masked_in_every_ledger_view(self):
        raw_phone = "1234567"
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-short-phone-0001", "short-phone.xlsx",
            xlsx([branch_row(phone=raw_phone)]),
        )
        rendered = json.dumps(preview, ensure_ascii=False)
        self.assertNotIn(raw_phone, rendered)
        masked_phone = preview["rows"][0]["data"]["manager_phone"]
        self.assertIn("*", masked_phone)
        self.assertNotEqual(raw_phone, re.sub(r"[^0-9]", "", masked_phone))

        fetched = inspectionimport.get_import(
            2, 20, preview["import_id"], "restaurant",
        )
        self.assertNotIn(raw_phone, json.dumps(fetched, ensure_ascii=False))
        staged = db.one(
            "SELECT masked_payload_json FROM inspection_branch_import_row "
            "WHERE import_id=?", (preview["import_id"],),
        )
        self.assertNotIn(raw_phone, staged["masked_payload_json"])

        committed = inspectionimport.commit_import(
            2, 20, preview["import_id"], "restaurant",
        )
        self.assertNotIn(raw_phone, json.dumps(committed, ensure_ascii=False))
        ledger_payloads = "\n".join(
            row["payload_json"] + row["masked_payload_json"]
            for row in db.q(
                "SELECT payload_json,masked_payload_json "
                "FROM inspection_branch_import_row WHERE import_id=?",
                (preview["import_id"],),
            )
        )
        self.assertNotIn(raw_phone, ledger_payloads)

    def test_duplicate_errors_block_atomic_commit_and_mask_pii(self):
        data = xlsx([branch_row("S001", "甲店"), branch_row("S001", "乙店", "13900139000")])
        preview = inspectionimport.preview_import(2, 20, "restaurant", "branch-dupe-0001", "dupe.xlsx", data)
        self.assertEqual(2, preview["counts"]["error"])
        self.assertTrue(all(row["error_code"] == "DUPLICATE_STORE_CODE" for row in preview["rows"]))
        self.assertNotIn("13900139000", json.dumps(preview, ensure_ascii=False))
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "IMPORT_HAS_ERRORS"):
            inspectionimport.commit_import(2, 20, preview["import_id"], "restaurant")
        self.assertEqual(0, db.one("SELECT COUNT(*) n FROM store_branch")["n"])

    def test_error_preview_never_stages_raw_pii_or_business_values(self):
        valid = branch_row("SAFE001", "混合有效店", "13987654321")
        valid[7] = "隐私经理甲"
        valid[8] = "PRIVATEEMP9001"
        invalid = branch_row("BAD002", "无效店", "13765432109")
        invalid[1] = ""  # one invalid row makes the whole preview uncommittable
        data = xlsx([valid, invalid], [[
            "SAFE001", "common.net_revenue", "2026-03-01", "2026-03-31",
            876543.125, "CNY", "private-source-991", "private-value-note-991",
        ]])

        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-pii-error-0001", "mixed.xlsx", data
        )
        self.assertEqual(1, preview["counts"]["error"])
        ledger = db.one(
            "SELECT business_values_json FROM inspection_branch_import "
            "WHERE id=?", (preview["import_id"],),
        )
        self.assertEqual("[]", ledger["business_values_json"])
        staged = db.q(
            "SELECT payload_json,masked_payload_json "
            "FROM inspection_branch_import_row WHERE import_id=? "
            "ORDER BY row_number",
            (preview["import_id"],),
        )
        self.assertTrue(staged)
        self.assertTrue(all(
            row["payload_json"] == row["masked_payload_json"]
            for row in staged
        ))

        # Scan the logical database, not only the public response.  An errored
        # preview can never commit, so no raw PII or operating value may enter
        # SQLite even temporarily through an insert-then-redact sequence.
        database_dump = "\n".join(db.conn().iterdump())
        for secret in (
            "隐私经理甲", "PRIVATEEMP9001", "13987654321",
            "876543.125", "private-source-991", "private-value-note-991",
        ):
            self.assertNotIn(secret, database_dump)
        with self.assertRaisesRegex(
            inspectionimport.ImportContractError, "IMPORT_HAS_ERRORS"
        ):
            inspectionimport.commit_import(
                2, 20, preview["import_id"], "restaurant"
            )
        self.assertEqual(0, db.one("SELECT COUNT(*) n FROM store_branch")["n"])
        self.assertEqual(
            0, db.one("SELECT COUNT(*) n FROM inspection_business_value")["n"]
        )

    def test_strict_isolated_parser_rejects_extension_formula_link_and_bomb(self):
        valid = xlsx([branch_row()])
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "FILE_TYPE_UNSUPPORTED"):
            inspectionimport.preview_import(2, 20, "restaurant", "branch-type-0001", "bad.xlsm", valid)
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "XLSX_FORMULA_FORBIDDEN"):
            inspectionimport.preview_import(2, 20, "restaurant", "branch-formula-0001", "bad.xlsx", xlsx([branch_row()], formula=True))
        linked = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(valid)) as source, zipfile.ZipFile(linked, "w") as target:
            for item in source.infolist():
                target.writestr(item, source.read(item.filename))
            target.writestr("xl/externalLinks/externalLink1.xml", "<externalLink/>")
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "XLSX_EXTERNAL_LINK_FORBIDDEN"):
            inspectionimport.preview_import(2, 20, "restaurant", "branch-link-0001", "bad.xlsx", linked.getvalue())
        # XML permits whitespace around '='. Security checks must parse the
        # relationship semantics instead of matching one serialization.
        spaced_link = io.BytesIO()
        relationship = (
            '<?xml version="1.0"?><Relationships '
            'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId99" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            'relationships/hyperlink" Target="https://attacker.invalid/track" '
            'TargetMode = "External"/></Relationships>'
        )
        with zipfile.ZipFile(io.BytesIO(valid)) as source, zipfile.ZipFile(
            spaced_link, "w", zipfile.ZIP_DEFLATED
        ) as target:
            for item in source.infolist():
                target.writestr(item, source.read(item.filename))
            target.writestr(
                "xl/worksheets/_rels/sheet1.xml.rels", relationship
            )
        with self.assertRaisesRegex(
            inspectionimport.ImportContractError,
            "XLSX_EXTERNAL_LINK_FORBIDDEN",
        ):
            inspectionimport.preview_import(
                2, 20, "restaurant", "branch-link-0002", "bad.xlsx",
                spaced_link.getvalue(),
            )
        bomb = io.BytesIO()
        with zipfile.ZipFile(bomb, "w", zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("xl/worksheets/sheet1.xml", b"0" * (2 * 1024 * 1024))
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "XLSX_ZIP_BOMB"):
            inspectionimport.preview_import(2, 20, "restaurant", "branch-bomb-0001", "bad.xlsx", bomb.getvalue())

    def test_archive_xml_preflight_has_component_and_relationship_node_budgets(self):
        valid = xlsx([branch_row()])
        many_relationships = (
            '<?xml version="1.0"?><Relationships '
            'xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            + "".join(
                f'<Relationship Id="rId{index}" Type="internal" Target="sheet.xml"/>'
                for index in range(5_000)
            )
            + "</Relationships>"
        ).encode()
        node_bomb = rewritten_xlsx(
            valid, {"xl/_rels/workbook.xml.rels": many_relationships},
        )
        with self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport._validate_archive(node_bomb)
        self.assertEqual("XLSX_XML_LIMIT_EXCEEDED", caught.exception.code)

        with mock.patch.object(
            inspectionimport, "MAX_RELATIONSHIP_XML_BYTES", 64,
        ):
            oversized = rewritten_xlsx(
                valid,
                {"xl/_rels/workbook.xml.rels": b"<Relationships>" + b"x" * 65},
            )
            with self.assertRaises(inspectionimport.ImportContractError) as caught:
                inspectionimport._validate_archive(oversized)
            self.assertEqual("XLSX_XML_LIMIT_EXCEEDED", caught.exception.code)

        with mock.patch.object(
            inspectionimport, "MAX_SHARED_STRINGS_XML_BYTES", 64,
        ):
            metadata_bomb = rewritten_xlsx(
                valid, {"xl/sharedStrings.xml": b"x" * 65},
            )
            with self.assertRaises(inspectionimport.ImportContractError) as caught:
                inspectionimport._validate_archive(metadata_bomb)
            self.assertEqual("XLSX_XML_LIMIT_EXCEEDED", caught.exception.code)

    def test_worker_rejects_single_cell_and_cumulative_text_exhaustion(self):
        too_long = branch_row()
        too_long[17] = "X" * 4_097
        with self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport._parse_isolated(xlsx([too_long]))
        self.assertEqual("CELL_TEXT_LIMIT_EXCEEDED", caught.exception.code)

        rows = []
        for index in range(2_050):
            row = branch_row(code=f"TEXT{index:05d}")
            row[17] = "Y" * 4_096
            rows.append(row)
        with self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport._parse_isolated(xlsx(rows))
        self.assertEqual("TEXT_BUDGET_EXCEEDED", caught.exception.code)

    def test_parent_kills_worker_as_soon_as_streamed_stdout_exceeds_cap(self):
        command = [
            sys.executable, "-I", "-c",
            "import sys,time;sys.stdout.buffer.write(b'x'*4096);"
            "sys.stdout.buffer.flush();time.sleep(10)",
        ]
        started = time.monotonic()
        with mock.patch.object(
            inspectionimport, "MAX_WORKER_STDOUT_BYTES", 1_024,
        ), self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport._run_worker_bounded(
                command, cwd=tempfile.gettempdir(), env={"PATH": os.environ.get("PATH", "")},
            )
        self.assertEqual("XLSX_INVALID", caught.exception.code)
        self.assertLess(time.monotonic() - started, 3.0)

    def test_real_download_template_parses_without_contract_drift(self):
        template = Path(__file__).parents[1] / "static" / "inspection-store-import-template.xlsx"
        data = template.read_bytes()
        inspectionimport._validate_archive(data)
        parsed = inspectionimport._parse_isolated(data)
        self.assertEqual([], parsed["branches"])
        self.assertEqual([], parsed["business_values"])

    def test_business_rows_use_catalog_units_sources_and_separate_row_kind(self):
        data = xlsx([branch_row()], [[
            "S001", "invented.metric", "2026-01-01", "2026-01-31",
            1, "widgets", "", "",
        ]])
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-metric-0001", "metric.xlsx", data
        )
        error = next(row for row in preview["rows"] if row["row_kind"] == "business")
        self.assertEqual(4, error["row_number"])
        self.assertEqual("BUSINESS_VALUE_INVALID", error["error_code"])
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "IMPORT_HAS_ERRORS"):
            inspectionimport.commit_import(
                2, 20, preview["import_id"], "restaurant"
            )

    def test_malformed_business_store_codes_become_pageable_error_rows(self):
        for position, bad_code in enumerate(("X" * 41, "BAD\x7fCODE"), 1):
            with self.subTest(bad_code=repr(bad_code)):
                preview = inspectionimport.preview_import(
                    2, 20, "restaurant", f"business-bad-code-{position:04d}",
                    f"bad-code-{position}.xlsx",
                    xlsx([branch_row()], [[
                        bad_code, "common.net_revenue", "2026-01-01",
                        "2026-01-31", 1, "CNY", "audited-pos", "",
                    ]]),
                )
                self.assertEqual(1, preview["business_counts"]["error"])
                page = inspectionimport.get_import(
                    2, 20, preview["import_id"], "restaurant",
                    row_kind="business", errors_only=True,
                )
                self.assertEqual(1, page["filtered_total_rows"])
                self.assertEqual("BUSINESS_VALUE_INVALID", page["rows"][0]["error_code"])
                safe_code = page["rows"][0]["store_code"] or ""
                self.assertLessEqual(len(safe_code), 40)
                self.assertIsNone(inspectionimport._CONTROL_RE.search(safe_code))

    def test_same_source_cannot_amplify_active_preview_and_stale_source_can_retry(self):
        data = xlsx([branch_row(phone="13712345678")])
        first = inspectionimport.preview_import(
            2, 20, "restaurant", "source-dedupe-first", "first.xlsx", data,
        )
        replay = inspectionimport.preview_import(
            2, 20, "restaurant", "source-dedupe-first", "first.xlsx", data,
        )
        self.assertEqual(first["import_id"], replay["import_id"])
        reused = inspectionimport.preview_import(
            2, 20, "restaurant", "source-dedupe-second", "copy.xlsx", data,
        )
        self.assertEqual(first["import_id"], reused["import_id"])
        self.assertTrue(reused["source_reused"])
        self.assertEqual("source-dedupe-second", reused["request_key"])
        self.assertNotEqual(first["request_key"], reused["request_key"])
        self.assertEqual(
            1,
            db.one("SELECT COUNT(*) n FROM inspection_branch_import")["n"],
        )

        db.execute(
            "UPDATE inspection_branch_import SET updated_at=? WHERE id=?",
            (time.time() - inspectionimport.PREVIEW_TTL_SECONDS - 1, first["import_id"]),
        )
        replacement = inspectionimport.preview_import(
            2, 20, "restaurant", "source-dedupe-third", "retry.xlsx", data,
        )
        self.assertNotEqual(first["import_id"], replacement["import_id"])
        self.assertEqual(
            "expired",
            db.one(
                "SELECT status FROM inspection_branch_import WHERE id=?",
                (first["import_id"],),
            )["status"],
        )

    def test_expired_preview_clears_raw_staging_and_keeps_masked_audit_summary(self):
        raw_phone = "13798765432"
        raw_employee = "PRIVATE-EMP-777"
        raw_source = "private-erp-source-777"
        row = branch_row(phone=raw_phone)
        row[8] = raw_employee
        expiry_data = xlsx([row], [[
            "S001", "common.net_revenue", "2026-04-01", "2026-04-30",
            987654.321, "CNY", raw_source, "private-business-note-777",
        ]])
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "preview-expiry-0001", "expiry.xlsx",
            expiry_data,
        )
        db.execute(
            "UPDATE inspection_branch_import SET updated_at=? WHERE id=?",
            (time.time() - inspectionimport.PREVIEW_TTL_SECONDS - 1, preview["import_id"]),
        )
        expired = inspectionimport.get_import(
            2, 20, preview["import_id"], "restaurant",
        )
        self.assertEqual("expired", expired["status"])
        self.assertEqual(1, expired["total_rows"])
        self.assertEqual(1, expired["business_total_rows"])
        ledger = db.one(
            "SELECT business_values_json,staging_purged_at "
            "FROM inspection_branch_import WHERE id=?",
            (preview["import_id"],),
        )
        self.assertEqual("[]", ledger["business_values_json"])
        self.assertIsNotNone(ledger["staging_purged_at"])
        logical_ledger = "\n".join(
            row["payload_json"] + row["masked_payload_json"]
            for row in db.q(
                "SELECT payload_json,masked_payload_json "
                "FROM inspection_branch_import_row WHERE import_id=?",
                (preview["import_id"],),
            )
        )
        for secret in (
            raw_phone, raw_employee, raw_source, "987654.321",
            "private-business-note-777",
        ):
            self.assertNotIn(secret, logical_ledger)
        with self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport.commit_import(
                2, 20, preview["import_id"], "restaurant",
            )
            # Model a retained full-capacity audit summary without constructing
            # another 540,000 physical rows in this focused unit test.
            db.execute(
                "UPDATE inspection_branch_import SET total_rows=20000,"
                "business_create_count=40000 WHERE id=?",
                (preview["import_id"],),
            )
        self.assertEqual("IMPORT_PREVIEW_EXPIRED", caught.exception.code)
        with self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport.preview_import(
                2, 20, "restaurant", "preview-expiry-0001", "expiry.xlsx",
                expiry_data,
            )
        self.assertEqual("IMPORT_PREVIEW_EXPIRED", caught.exception.code)

    def test_clean_preview_never_writes_raw_staging_secrets_to_sqlite_or_wal(self):
        secrets = (
            "13765439876", "PHYSPIIEMPQZ917263",
            "private-source-qz917263", "private-note-qz917263", "777.125",
        )
        row = branch_row("CIPHER001", "密文店", secrets[0])
        row[8] = secrets[1]
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "cipher-at-rest-0001", "cipher.xlsx",
            xlsx([row], [[
                "CIPHER001", "common.net_revenue", "2026-07-01",
                "2026-07-31", 777.125, "CNY", secrets[2], secrets[3],
            ]]),
        )
        self.assertEqual("previewed", preview["status"])
        self.assertEqual(777.125, next(
            item["data"]["value"] for item in preview["rows"]
            if item["row_kind"] == "business"
        ))
        on_disk = b""
        for suffix in ("", "-wal"):
            path = Path(str(db.DB_PATH) + suffix)
            if path.exists():
                on_disk += path.read_bytes()
        for secret in secrets:
            self.assertNotIn(secret.encode("utf-8"), on_disk)

        committed = inspectionimport.commit_import(
            2, 20, preview["import_id"], "restaurant",
        )
        self.assertEqual("committed", committed["status"])
        self.assertEqual(
            secrets[0],
            db.one(
                "SELECT manager_phone FROM store_branch WHERE store_code='CIPHER001'"
            )["manager_phone"],
        )

    def test_staging_key_change_and_row_cipher_swap_fail_closed(self):
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "cipher-key-change-0001", "key.xlsx",
            xlsx([branch_row("KEY001", "密钥店")]),
        )
        with mock.patch.object(auth, "_secret", return_value=b"K" * 32):
            with self.assertRaises(inspectionimport.ImportContractError) as caught:
                inspectionimport.commit_import(
                    2, 20, preview["import_id"], "restaurant",
                )
        self.assertEqual("IMPORT_PREVIEW_EXPIRED", caught.exception.code)
        self.assertEqual(
            "expired",
            db.one(
                "SELECT status FROM inspection_branch_import WHERE id=?",
                (preview["import_id"],),
            )["status"],
        )

        swapped = inspectionimport.preview_import(
            2, 20, "restaurant", "cipher-row-swap-0001", "swap.xlsx",
            xlsx([
                branch_row("SWAP001", "密文一"),
                branch_row("SWAP002", "密文二", "13900139000"),
            ]),
        )
        rows = db.q(
            "SELECT row_number,payload_json FROM inspection_branch_import_row "
            "WHERE import_id=? ORDER BY row_number", (swapped["import_id"],),
        )
        db.execute(
            "UPDATE inspection_branch_import_row SET payload_json=? "
            "WHERE import_id=? AND row_number=?",
            (rows[0]["payload_json"], swapped["import_id"], rows[1]["row_number"]),
        )
        with self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport.commit_import(
                2, 20, swapped["import_id"], "restaurant",
            )
        self.assertEqual("IMPORT_PREVIEW_EXPIRED", caught.exception.code)
        self.assertEqual(0, db.one(
            "SELECT COUNT(*) n FROM store_branch WHERE store_code LIKE 'SWAP%'"
        )["n"])

    def test_active_preview_count_row_and_byte_quotas_are_atomic(self):
        first = inspectionimport.preview_import(
            2, 20, "restaurant", "quota-first-0001", "first.xlsx",
            xlsx([branch_row("QUOTA001")]),
        )
        with mock.patch.object(
            inspectionimport, "MAX_ACTIVE_PREVIEWS_PER_TENANT", 1,
        ), self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport.preview_import(
                2, 20, "restaurant", "quota-count-0002", "count.xlsx",
                xlsx([branch_row("QUOTA002")]),
            )
        self.assertEqual("IMPORT_PREVIEW_QUOTA_EXCEEDED", caught.exception.code)

        db.execute(
            "UPDATE inspection_branch_import SET updated_at=? WHERE id=?",
            (time.time() - inspectionimport.PREVIEW_TTL_SECONDS - 1, first["import_id"]),
        )
        inspectionimport.get_import(2, 20, first["import_id"], "restaurant")
        with mock.patch.object(
            inspectionimport, "MAX_ACTIVE_PREVIEW_ROWS_PER_TENANT", 1,
        ), self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport.preview_import(
                2, 20, "restaurant", "quota-rows-0003", "rows.xlsx",
                xlsx([branch_row("QUOTA003"), branch_row("QUOTA004")]),
            )
        self.assertEqual("IMPORT_PREVIEW_QUOTA_EXCEEDED", caught.exception.code)

        with mock.patch.object(
            inspectionimport, "MAX_ACTIVE_PREVIEW_BYTES_PER_TENANT", 1,
        ), self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport.preview_import(
                2, 20, "restaurant", "quota-bytes-0004", "bytes.xlsx",
                xlsx([branch_row("QUOTA005")]),
            )
        self.assertEqual("IMPORT_PREVIEW_QUOTA_EXCEEDED", caught.exception.code)
        self.assertEqual(
            0,
            db.one(
                "SELECT COUNT(*) n FROM inspection_branch_import "
                "WHERE status='previewed'",
            )["n"],
        )

    def test_retired_preview_quota_is_separate_from_active_recovery(self):
        data = xlsx([branch_row("QUOTACAS", "配额CAS店")])
        first = inspectionimport.preview_import(
            2, 20, "restaurant", "quota-cas-first-0001", "first.xlsx", data,
        )
        # Simulate the post-rollback CAS retirement path.  Even with an active
        # cap of one, the encrypted retired generation must not prevent the
        # owner from creating a fresh generation immediately.
        db.execute(
            "UPDATE inspection_branch_import SET status='expired' WHERE id=?",
            (first["import_id"],),
        )
        with mock.patch.object(
            inspectionimport, "MAX_ACTIVE_PREVIEWS_PER_TENANT", 1,
        ):
            replacement = inspectionimport.preview_import(
                2, 20, "restaurant", "quota-cas-second-0001", "second.xlsx",
                data,
            )
        self.assertNotEqual(first["import_id"], replacement["import_id"])

        inspectionimport.commit_import(
            2, 20, replacement["import_id"], "restaurant",
        )
        with mock.patch.object(
            inspectionimport, "MAX_RETIRED_PREVIEWS_PER_TENANT", 1,
        ), self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport.preview_import(
                2, 20, "restaurant", "quota-retired-third-0001", "third.xlsx",
                xlsx([branch_row("QUOTANEW", "失效代数配额店")]),
            )
        self.assertEqual("IMPORT_PREVIEW_QUOTA_EXCEEDED", caught.exception.code)

    def test_committed_history_never_blocks_future_preview(self):
        for index in range(9):
            preview = inspectionimport.preview_import(
                2, 20, "restaurant", f"history-commit-{index:04d}",
                f"history-{index}.xlsx",
                xlsx([branch_row(f"HISTORY{index:03d}", f"历史店{index}")]),
            )
            inspectionimport.commit_import(
                2, 20, preview["import_id"], "restaurant",
            )
        tenth = inspectionimport.preview_import(
            2, 20, "restaurant", "history-preview-0010", "history-10.xlsx",
            xlsx([branch_row("HISTORY010", "历史店10")]),
        )
        self.assertEqual("previewed", tenth["status"])

    def test_actor_and_industry_scope_are_rechecked_for_ledger_reads(self):
        data = xlsx([branch_row()])
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "branch-scope-0001", "scope.xlsx", data
        )
        fetched = inspectionimport.get_import(
            2, 20, preview["import_id"], "restaurant"
        )
        self.assertEqual("restaurant", fetched["industry_key"])
        db.execute(
            "INSERT INTO tenant_industry(tenant_id,industry_key,is_primary,created_at) "
            "VALUES(2,'auto',0,0)"
        )
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "IMPORT_NOT_FOUND"):
            inspectionimport.get_import(2, 20, preview["import_id"], "auto")
        db.execute("UPDATE users SET enabled=0 WHERE id=20")
        with self.assertRaisesRegex(inspectionimport.ImportContractError, "SCOPE_FORBIDDEN"):
            inspectionimport.get_import(
                2, 20, preview["import_id"], "restaurant"
            )

    def test_preview_freezes_catalog_and_commit_rejects_version_or_hash_drift(self):
        data = xlsx([branch_row()], [[
            "S001", "common.net_revenue", "2026-01-01", "2026-01-31",
            1234.5, "CNY", "audited-pos", "",
        ]])
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "catalog-freeze-0001", "catalog.xlsx", data
        )
        original = inspectionstandards.version_summary("restaurant")
        self.assertEqual(original["catalog_version"], preview["catalog_version"])
        self.assertEqual(original["sha256"], preview["catalog_sha256"])

        drifted = {
            **original,
            "sha256": "f" * 64 if original["sha256"] != "f" * 64 else "e" * 64,
        }
        with mock.patch.object(
            inspectionstandards, "version_summary", return_value=drifted
        ), self.assertRaisesRegex(
            inspectionimport.ImportContractError, "IMPORT_STATE_CONFLICT"
        ):
            inspectionimport.commit_import(
                2, 20, preview["import_id"], "restaurant"
            )
        self.assertEqual(
            "expired",
            db.one(
                "SELECT status FROM inspection_branch_import WHERE id=?",
                (preview["import_id"],),
            )["status"],
        )
        self.assertEqual(0, db.one("SELECT COUNT(*) n FROM store_branch")["n"])
        self.assertEqual(
            0,
            db.one("SELECT COUNT(*) n FROM inspection_business_value")["n"],
        )

    def test_bulk_master_import_is_owner_only(self):
        db.insert("users", {
            "id": 21, "tenant_id": 2, "username": "regional-import",
            "password_hash": "test", "role": "member",
            "modules_json": '["restaurant"]', "enabled": 1,
        })
        with self.assertRaisesRegex(
            inspectionimport.ImportContractError, "SCOPE_FORBIDDEN"
        ):
            inspectionimport.preview_import(
                2, 21, "restaurant", "branch-member-0001",
                "branches.xlsx", xlsx([branch_row()]),
            )

    def test_business_comparison_uses_latest_real_values_and_never_invents(self):
        branch_id = db.insert("store_branch", {
            "tenant_id": 2, "industry_key": "restaurant",
            "store_code": "S001", "name": "一号店",
        })
        for start, end, value in (
            ("2025-01-01", "2025-01-31", 900),
            ("2025-12-01", "2025-12-31", 1100),
            ("2026-01-01", "2026-01-31", 1234.5),
        ):
            db.insert("inspection_business_value", {
                "tenant_id": 2, "industry_key": "restaurant",
                "branch_id": branch_id, "import_id": 1,
                "metric_key": "common.net_revenue", "period_start": start,
                "period_end": end, "value": value, "unit": "CNY",
                "source_ref": "erp",
            })
        result = inspectionimport.business_comparison(
            2, "restaurant", branch_id
        )
        revenue = next(
            item for item in result["metrics"]
            if item["metric_code"] == "common.net_revenue"
        )
        self.assertEqual(1234.5, revenue["actual"])
        self.assertEqual(1100, revenue["previous_period"])
        self.assertEqual(900, revenue["same_period_last_year"])
        self.assertIsNone(revenue["target"])
        self.assertIsNone(revenue["benchmark"])
        definition = next(
            item for item in inspectionstandards.metric_catalog("restaurant")
            if item["metric_code"] == "common.net_revenue"
        )
        self.assertEqual(definition["formula"], revenue["formula"])
        self.assertEqual(
            definition["required_inputs"], revenue["required_inputs"]
        )
        self.assertIsNone(revenue["previous_period_reason_code"])
        self.assertIsNone(revenue["same_period_last_year_reason_code"])
        self.assertEqual("target_not_configured", revenue["target_reason_code"])
        self.assertEqual(
            "benchmark_not_configured", revenue["benchmark_reason_code"]
        )
        self.assertEqual({
            "actual": None,
            "previous_period": None,
            "same_period_last_year": None,
            "target": "target_not_configured",
            "benchmark": "benchmark_not_configured",
        }, revenue["reason_codes"])
        missing = next(
            item for item in result["metrics"]
            if item["metric_code"] == "common.employee_count"
        )
        self.assertFalse(missing["availability"])
        self.assertIsNone(missing["actual"])
        self.assertEqual(
            "metric_data_unavailable", missing["actual_reason_code"]
        )
        self.assertEqual(
            "previous_period_unavailable",
            missing["previous_period_reason_code"],
        )
        self.assertEqual(
            "same_period_last_year_unavailable",
            missing["same_period_last_year_reason_code"],
        )


if __name__ == "__main__":
    unittest.main()
