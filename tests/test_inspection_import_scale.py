"""Scale and pagination contracts for the schema-52 branch XLSX importer."""
from __future__ import annotations

import io
import json
import os
import tempfile
import unittest

from openpyxl import Workbook

from app import db, inspectionimport, main


BRANCH_HEADERS = [
    "门店编号*", "店名*", "区域*", "省", "市", "区/县", "详细地址*", "店长姓名", "店长工号",
    "店长手机号", "门店类型", "开业日期", "营业面积㎡", "座位数/房间数/工位数",
    "经度", "纬度", "启用状态*", "备注",
]
VALUE_HEADERS = [
    "门店编号*", "指标编码*", "期间开始*", "期间结束*", "数值*",
    "单位*", "数据来源*", "备注",
]


def bulk_xlsx(
    branch_count: int,
    *,
    business_count: int = 0,
    duplicate_code: bool = False,
) -> bytes:
    """Build a low-memory, contract-exact workbook for scale tests."""
    book = Workbook(write_only=True)
    branches = book.create_sheet("门店主表")
    branches.append(["派活 · 门店批量导入模板"])
    branches.append(["填写说明"])
    branches.append(BRANCH_HEADERS)
    for index in range(1, branch_count + 1):
        code = "DUPLICATE" if duplicate_code else f"S{index:05d}"
        branches.append([
            code, f"规模门店{index}", f"区域{index % 20:02d}", "浙江省",
            "杭州市", "西湖区", f"文三路{index}号", f"店长{index}",
            f"E{index:06d}", "13800138000", "标准店", "2026-01-02",
            120.5, 60, 120.123, 30.234, "启用", "规模测试",
        ])

    values = book.create_sheet("经营数据")
    values.append(["派活 · 门店经营数据（选填）"])
    values.append(["填写说明"])
    values.append(VALUE_HEADERS)
    for index in range(1, business_count + 1):
        store_index = ((index - 1) % max(1, branch_count)) + 1
        period_index = ((index - 1) // max(1, branch_count)) + 1
        month = min(period_index, 12)
        code = "DUPLICATE" if duplicate_code else f"S{store_index:05d}"
        values.append([
            code, "common.net_revenue", f"2026-{month:02d}-01",
            f"2026-{month:02d}-28",
            float(index), "CNY", "audited-pos", "",
        ])

    book.create_sheet("填写说明")
    book.create_sheet("示例（不要导入）")
    output = io.BytesIO()
    book.save(output)
    return output.getvalue()


class InspectionImportScaleCase(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.old_path = db.DB_PATH
        db._close_all_connections()
        db._conn = db._conn_path = None
        db.DB_PATH = os.path.join(self.tmp.name, "scale.db")
        db.conn()
        db.insert("tenants", {"id": 2, "name": "规模企业"})
        db.execute(
            "INSERT INTO tenant_industry(tenant_id,industry_key,is_primary,created_at) "
            "VALUES(2,'restaurant',1,0)"
        )
        db.insert("users", {
            "id": 20, "tenant_id": 2, "username": "scale-owner",
            "password_hash": "test", "role": "owner", "modules_json": "[]",
            "enabled": 1,
        })

    def tearDown(self):
        db._close_all_connections()
        db._conn = db._conn_path = None
        db.DB_PATH = self.old_path
        self.tmp.cleanup()

    def test_isolated_worker_parses_ten_thousand_standard_branches(self):
        data = bulk_xlsx(10_000)
        inspectionimport._validate_archive(data)
        parsed = inspectionimport._parse_isolated(data)
        self.assertEqual(10_000, len(parsed["branches"]))
        self.assertEqual("S00001", parsed["branches"][0]["data"]["门店编号*"])
        self.assertEqual("S10000", parsed["branches"][-1]["data"]["门店编号*"])

    def test_isolated_worker_preserves_full_twenty_thousand_by_forty_thousand_capacity(self):
        data = bulk_xlsx(20_000, business_count=40_000)
        inspectionimport._validate_archive(data)
        parsed = inspectionimport._parse_isolated(data)
        self.assertEqual(20_000, len(parsed["branches"]))
        self.assertEqual(40_000, len(parsed["business_values"]))
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "scale-preview-20000-40000",
            "full-capacity.xlsx", data,
        )
        self.assertEqual(20_000, preview["counts"]["create"])
        self.assertEqual(40_000, preview["business_counts"]["create"])

    def test_twenty_thousand_preview_is_bounded_then_commits_and_searches(self):
        data = bulk_xlsx(20_000, business_count=20_000)
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "scale-preview-20000", "all-stores.xlsx", data,
        )
        self.assertEqual(20_000, preview["total_rows"])
        self.assertEqual(20_000, preview["business_total_rows"])
        self.assertEqual(20_000, preview["counts"]["create"])
        self.assertEqual(20_000, preview["business_counts"]["create"])
        self.assertEqual(50, len(preview["rows"]))
        self.assertTrue(preview["has_more"])
        self.assertEqual(40_000, preview["filtered_total_rows"])
        self.assertIsNotNone(preview["next_cursor"])

        second = inspectionimport.get_import(
            2, 20, preview["import_id"], "restaurant",
            limit=200, cursor=preview["next_cursor"], row_kind="branch",
        )
        self.assertEqual(200, len(second["rows"]))
        self.assertEqual(54, second["rows"][0]["row_number"])
        business = inspectionimport.get_import(
            2, 20, preview["import_id"], "restaurant",
            limit=25, row_kind="business",
        )
        self.assertEqual(20_000, business["filtered_total_rows"])
        self.assertEqual(25, len(business["rows"]))
        self.assertTrue(all(
            row["row_kind"] == "business" for row in business["rows"]
        ))
        self.assertEqual(4, business["rows"][0]["row_number"])

        committed = inspectionimport.commit_import(
            2, 20, preview["import_id"], "restaurant",
        )
        self.assertEqual("committed", committed["status"])
        self.assertEqual(50, len(committed["rows"]))
        self.assertEqual(
            20_000, db.one("SELECT COUNT(*) n FROM store_branch")["n"],
        )
        self.assertEqual(
            20_000,
            db.one("SELECT COUNT(*) n FROM inspection_business_value")["n"],
        )
        found = main._inspection_branch_search_db(
            2, 20, "restaurant", q="S19999", limit=20,
        )
        self.assertEqual(["S19999"], [row["store_code"] for row in found["items"]])

    def test_error_rows_are_masked_and_independently_pageable(self):
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "scale-errors-page", "errors.xlsx",
            bulk_xlsx(260, duplicate_code=True),
        )
        self.assertEqual(260, preview["counts"]["error"])
        self.assertEqual(50, len(preview["rows"]))
        first = inspectionimport.get_import(
            2, 20, preview["import_id"], "restaurant",
            limit=200, errors_only=True,
        )
        self.assertEqual(260, first["filtered_total_rows"])
        self.assertEqual(200, len(first["rows"]))
        self.assertTrue(first["has_more"])
        second = inspectionimport.get_import(
            2, 20, preview["import_id"], "restaurant",
            limit=200, cursor=first["next_cursor"], errors_only=True,
        )
        self.assertEqual(60, len(second["rows"]))
        self.assertFalse(second["has_more"])
        rendered = json.dumps([first, second], ensure_ascii=False)
        self.assertNotIn("13800138000", rendered)
        self.assertIn("138****8000", rendered)
        self.assertTrue(all(
            row["error_code"] == "DUPLICATE_STORE_CODE"
            for page in (first, second) for row in page["rows"]
        ))

    def test_branch_limit_overflow_fails_with_stable_code(self):
        with self.assertRaises(inspectionimport.ImportContractError) as caught:
            inspectionimport.preview_import(
                2, 20, "restaurant", "scale-over-limit", "too-many.xlsx",
                bulk_xlsx(20_001),
            )
        self.assertEqual("ROW_LIMIT_EXCEEDED", caught.exception.code)

    def test_page_contract_rejects_unbounded_or_malformed_requests(self):
        preview = inspectionimport.preview_import(
            2, 20, "restaurant", "scale-invalid-page", "page.xlsx",
            bulk_xlsx(1),
        )
        for kwargs in (
            {"limit": 201},
            {"limit": 0},
            {"cursor": "-1"},
            {"cursor": "not-a-cursor"},
            {"row_kind": "secret"},
        ):
            with self.subTest(kwargs=kwargs), self.assertRaises(
                inspectionimport.ImportContractError,
            ) as caught:
                inspectionimport.get_import(
                    2, 20, preview["import_id"], "restaurant", **kwargs,
                )
            self.assertEqual("IMPORT_PAGE_INVALID", caught.exception.code)


if __name__ == "__main__":
    unittest.main()
